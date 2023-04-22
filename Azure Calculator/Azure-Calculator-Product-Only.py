import os
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
import openpyxl

# Set constants for this script
EXCEL_FILE_PATH_HERE = 'C:/Users/tluk/Documents/Scripts/Python/Web Browser Automation/Product-Only.xlsx'
WHERE_SELENIUM_CHROME_DRIVE_IS_LOCATED = 'C:/SeleniumDrivers'
URL_TO_AZURE_CALCULATOR = 'https://azure.microsoft.com/en-ca/pricing/calculator/'
ITEMS_TO_IGNORE = ["RedPanda", "Power BI (Included on m365)", "Power BI Premium"]
ITEMS_TO_IGNORE_REGION = ["Azure DevOps", "Azure Policy", "Virtual Network", "Azure DDoS Protection"]

# Set the path to the Excel file containing the virtual machines to be configured
excel_file_path = EXCEL_FILE_PATH_HERE

# Use openpyxl to open the Excel file and select the active worksheet
workbook = openpyxl.load_workbook(excel_file_path)
worksheet = workbook.active

# Get the amount of rows and set the current row at 1 so we can stop at the end (openpyxl doesn't maintain a row index)
row_count = worksheet.max_row
current_row = 1

# Start the Chrome webdriver (where chromedriver.exe is accessed)
os.environ['PATH'] += WHERE_SELENIUM_CHROME_DRIVE_IS_LOCATED

# Use options with detach to stop the driver from closing the browser so the estimate can be saved
chrome_options = Options()
chrome_options.add_experimental_option("detach", True)
driver = webdriver.Chrome(options=chrome_options)

# Launch Google Chrome and navigate to the Azure calculator website
driver.get(URL_TO_AZURE_CALCULATOR)

# Maximize the Google Chrome window
driver.maximize_window()

# Wait for the survey message popup to load
driver.implicitly_wait(5)

# Handle the survey message popup:
'''
Microsoft is conducting an online survey to understand your opinions about the Microsoft Azure website. If you choose to participate, the online survey will be presented to you when you leave the website.

Would you like to participate?
'''
try:
    survey_button = usage_computeBillingOption_input = driver.find_element(By.XPATH, '//input[@type="button" and @onclick="return COMSCORE.SiteRecruit.Builder.invitation.decline(this); return false;"]')
    survey_button.click()
except:
    pass

# Wait for the Azure Calculator page to load
driver.implicitly_wait(10)

# Loop through each row in the Excel worksheet and enter the data for each product into the Azure calculator
for row in worksheet.iter_rows(min_row=1): 
    if current_row == 1:
        estimate_name = row[0].value # grabbing the estimate name
        
        # Increment the current row by 1 now that the current row has been processed
        current_row += 1
    elif current_row == 2:
        # Skip the row because this containers header information
        # Increment the current row by 1 now that the current row has been processed
        current_row += 1

    elif current_row == 3: # First row of product data to be proccssed (assume this product has a region)
        item_name = row[0].value # The product name
        item_custom_name = row[1].value # What we're naming the first item
        region = row[2].value # The region for this product

        # Locate the product search field at the top of the page
        search_field = driver.find_element(By.XPATH, '//input[@class="product-search" and @aria-label="Search products" and @placeholder="Search products"]') 
        # Use the search field xpath and scroll back up in the browser
        actions = ActionChains(driver)
        actions.move_to_element(search_field).perform()
        # Clear the search field
        search_field.clear()
        # Search for the product
        search_field.send_keys(item_name)
        # Locate the button to add virtual machine
        add_item_button = driver.find_element(By.XPATH, '//button[@title="' + item_name + '"]') 
        # Click on the item_name button to add a new one
        add_item_button.click()

        # Select the region
        region_select = Select(driver.find_element(By.XPATH, '//select[@aria-label="Region" and @name="region"]'))
        region_select.select_by_visible_text(region)
       
        # Name item custom name entry
        item_custom_name_input = driver.find_element(By.XPATH, '//input[@placeholder="' + item_name + '"]')
        item_custom_name_input.clear()
        item_custom_name_input.send_keys(item_custom_name)
        
        # Minimize the configured product in preparation for the next product (if we do not minimize the heading then we will overwrite the first product with the next product configuration)
        # Try to find the product heading that is expanded and if found, click on it
        try:
            item_heading = driver.find_element(By.XPATH, '//span[contains(text(),"' + item_name + '") and @aria-expanded="true"]')
            # Find the button and scroll back up in the browser
            actions = ActionChains(driver)
            actions.move_to_element(item_heading).perform()
            item_heading.click()
        # Do not click on the product heading if it is not expanded (the top xpath doesn't exist)
        except:
            pass

        # Name Estimate
        estimate_name = estimate_name
        estimate_name_field = driver.find_element(By.XPATH, '//input[@id="estimate-name"]')
        estimate_name_field.clear()
        estimate_name_field.send_keys(estimate_name)

        # Increment the current row by 1 now that the current row has been processed
        current_row += 1
    elif current_row <= row_count: # Handle the rest of the rows
        item_name = row[0].value
        item_custom_name = row[1].value
        region = row[2].value

        # Exclude items not in Azure product catalog
        if item_name not in ITEMS_TO_IGNORE:
            # Locate the product search field at the top of the page
            search_field = driver.find_element(By.XPATH, '//input[@class="product-search" and @aria-label="Search products" and @placeholder="Search products"]') 
            # Use the search field xpath and scroll back up in the browser
            actions = ActionChains(driver)
            actions.move_to_element(search_field).perform()
            # Clear the search field
            search_field.clear()
            # Search for the product
            search_field.send_keys(item_name)
            # Locate the button to add virtual machine
            add_item_button = driver.find_element(By.XPATH, '//button[@title="' + item_name + '" and @type="submit"]') 
            # Click on the item_name button to add a new one
            add_item_button.click()
            # Skip region select for products without region
            if item_name not in ITEMS_TO_IGNORE_REGION:
                # Select the region
                region_select = Select(driver.find_element(By.XPATH, '//select[@name="region"]'))
                region_select.select_by_visible_text(region)

            # Name item custom name entry
            item_custom_name_input = driver.find_element(By.XPATH, '//input[@placeholder="' + item_name + '"]')
            item_custom_name_input.clear()
            item_custom_name_input.send_keys(item_custom_name)

            # Minimize the configured product in preparation for the next product (if we do not minimize the heading then we will overwrite the first product with the next product configuration)
            # Try to find the product heading that is expanded and if found, click on it
            try:
                item_heading = driver.find_element(By.XPATH, '//span[contains(text(),"' + item_name + '") and @aria-expanded="true"]')
                # Find the button and scroll back up in the browser
                actions_minimize = ActionChains(driver)
                actions_minimize.move_to_element(item_heading).perform()
                driver.execute_script("scrollBy(0,-100);") # Adjust for banner that overlays/covers the product heading we want to minimize
                item_heading.click()
            # Do not click on the product heading if it is not expanded (the top xpath doesn't exist)
            except:
                pass

            # Increment the current row by 1 now that the current row has been processed
            current_row += 1

# The commented out lines below are used to prevent Chrome Driver from closing the browser
# Close the webdriver
# driver.quit()
