'''
This python script will generate an estimate on the Azure Calculator with a list of virtual machines in an Excel spreadsheet. 
Please refer to my blog post for more information: http://terenceluk.blogspot.com/2023/04/automating-creation-of-azure-calculator.html
'''

import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
import openpyxl

# Set constants for this script
EXCEL_FILE_PATH_HERE = 'C:/Users/tluk/Documents/Scripts/Python/Web Browser Automation/VMs-Extended.xlsx'
WHERE_SELENIUM_CHROME_DRIVE_IS_LOCATED = 'C:/SeleniumDrivers'
URL_TO_AZURE_CALCULATOR = 'https://azure.microsoft.com/en-ca/pricing/calculator/'
AZURE_CALCULATOR_ESTIMATE_NAME = "Terence's New Estimate"

# Set the path to the Excel file containing the virtual machines to be configured
excel_file_path = EXCEL_FILE_PATH_HERE

# Use openpyxl to open the Excel file and select the active worksheet
workbook = openpyxl.load_workbook(excel_file_path)
worksheet = workbook.active

# Get the amount of rows and set the current row at 1 so we can stop at the end (openpyxl doesn't maintain a row index)
row_count = worksheet.max_row
current_row = 1

# Start the Chrome webdriver (where chromedriver.exe is accessed)
os.environ['PATH'] += WHERE_SELENIUM_CHROME_DRIVE_IS_LOCATED

# Use options with detach to stop the driver from closing the browser so the estimate can be saved
chrome_options = Options()
chrome_options.add_experimental_option("detach", True)
driver = webdriver.Chrome(options=chrome_options)

# Launch Google Chrome and navigate to the Azure calculator website
driver.get(URL_TO_AZURE_CALCULATOR)

# Maximize the Google Chrome window
driver.maximize_window()

# Wait for the survey message popup to load
driver.implicitly_wait(5)

# Handle the survey message popup:
'''
Microsoft is conducting an online survey to understand your opinions about the Microsoft Azure website. If you choose to participate, the online survey will be presented to you when you leave the website.

Would you like to participate?
'''
try:
    survey_button = usage_computeBillingOption_input = driver.find_element(By.XPATH, '//input[@type="button" and @onclick="return COMSCORE.SiteRecruit.Builder.invitation.decline(this); return false;"]')
    survey_button.click()
except:
    pass

# Wait for the Azure Calculator page to load
driver.implicitly_wait(10)

# Click on the "Add Virtual Machines" button to add the first VM so we can name the estimate (we can't name the estimate without a resource)
add_vm_button = driver.find_element(By.XPATH, '//button[@title="Virtual Machines"]')
add_vm_button.click()

# Name Estimate
estimate_name = AZURE_CALCULATOR_ESTIMATE_NAME
estimate_name_field = driver.find_element(By.XPATH, '//input[@id="estimate-name"]')
estimate_name_field.clear()
estimate_name_field.send_keys(estimate_name)

# Loop through each row in the Excel worksheet and enter the data for each VM into the Azure calculator
for row in worksheet.iter_rows(min_row=2):
    vm_name = row[0].value
    region = row[1].value
    operating_system = row[2].value
    os_type = row[3].value
    tier = row[4].value
    vm_size = row[5].value
    quantity = row[6].value
    usage_hours = row[7].value
    computeBillingOption = row[8].value
    managed_disk_tier = row[9].value
    managed_disk_size = row[10].value
    managed_disk_qty = row[11].value
    
    # Name virtual machine entry
    vm_name_input = driver.find_element(By.XPATH, '//input[@placeholder="Virtual Machines"]')
    vm_name_input.clear()
    vm_name_input.send_keys(vm_name)

    # Select the region
    region_select = Select(driver.find_element(By.XPATH, '//select[@name="region"]'))
    region_select.select_by_visible_text(region)

    # Select the Operating System
    operating_system_select = Select(driver.find_element(By.XPATH, '//select[@name="operatingSystem"]'))
    operating_system_select.select_by_visible_text(operating_system)

    # Select the Operating System Type
    os_type_select = Select(driver.find_element(By.XPATH, '//select[@name="type"]'))
    os_type_select.select_by_visible_text(os_type)

    # Select the Tier
    tier_select = Select(driver.find_element(By.XPATH, '//select[@name="tier"]'))
    tier_select.select_by_visible_text(tier)

    # Enter the VM Instance (vm_size)
    vm_size_input = driver.find_element(By.XPATH, '//input[@class="instancesSearchDropdown__input" and @role="combobox"]')
    # Send the VM size string (doesn't need to be full) then hit enter
    vm_size_input.send_keys(vm_size + "\n")

    # Enter the quantity (amount of these VMs)
    quantity_input = driver.find_element(By.XPATH, '//input[@name="count"]')
    quantity_input.clear()
    quantity_input.send_keys(quantity)

    # Select Compute Billing Options
    '''
    Options: 
        payg
        sv-one-year
        sv-three-year
        one-year
        three-year
    '''
    usage_computeBillingOption_input = driver.find_element(By.XPATH, '//input[@data-name-override="computeBillingOption" and @value="' + computeBillingOption + '"]')
    usage_computeBillingOption_input.click()

    # Click on and expand Managed Disks Heading
    # Try to find the Managed Disks heading that is not expanded and if found, click on it (this is so we can populate the parameters or we will get a "element not found")
    try:
        managed_disks_heading = driver.find_element(By.XPATH, '//button[text()="Managed Disks" and @aria-expanded="false"]')
        managed_disks_heading.click()
    # Do not click on the Managed Disks heading if it is already expanded (determined by the top xpath doesn't exist)
    except:
        pass

    # Enter Managed Disks Tier
    managed_disks_tier_input = Select(driver.find_element(By.XPATH, '//select[@aria-label="Tier" and @name="managedDiskTier"]'))
    managed_disks_tier_input.select_by_visible_text(managed_disk_tier)

    # Enter Managed Disks Size
    managed_disks_size_input = Select(driver.find_element(By.XPATH, '//select[@aria-label="Disk size" and @name="managedDiskType"]'))
    managed_disks_size_input.select_by_value(managed_disk_size)

    # Enter Managed Disks QTY
    managed_disk_qty_input = driver.find_element(By.XPATH, '//input[@aria-label="Disks" or @aria-label="Disk" and @name="managedDisks"]')
    managed_disk_qty_input.clear()
    managed_disk_qty_input.send_keys(managed_disk_qty)

    # Minimize the configured VM in preparation for the next VM (if we do not minimize the heading then we will overwrite the first VM with the next VM configuration)
    # Try to find the VM heading that is expanded and if found, click on it
    try:
        vm_heading = driver.find_element(By.XPATH, '//span[contains(text(),"' + vm_name + '") and @aria-expanded="true"]')
        # Find the button and scroll back up in the browser
        actions = ActionChains(driver)
        actions.move_to_element(vm_heading).perform()
        vm_heading.click()
    # Do not click on the VM heading if it is not expanded (the top xpath doesn't exist)
    except:
        pass
    
    # Increment the current row by 1 now that the current row has been processed
    current_row += 1
    
    # Check to see if we are at the end of the worksheet by comparing current row and the count of the rows
    if current_row < row_count:
        # If there are still VMs to be added then...
        # Locate the product search field at the top of the page
        search_field = driver.find_element(By.XPATH, '//input[@class="product-search" and @aria-label="Search products" and @placeholder="Search products"]') 
        # Use the search field xpath and scroll back up in the browser
        actions = ActionChains(driver)
        actions.move_to_element(search_field).perform()
        # Clear the search field (there would be an entry if we are on the 3rd VM)
        search_field.clear()
        # Search for virtual machines to eliminate //button[@title="Virtual Machines"] xpath showing up twice
        search_field.send_keys("Virtual machines")
        # Locate the button to add virtual machine
        add_vm_button = driver.find_element(By.XPATH, '//button[@title="Virtual Machines"]') 
        # Click on the "Add Virtual Machines" button to add a new one
        add_vm_button.click()

# The commented out lines below are used to prevent Chrome Driver from closing the browser
# Close the webdriver
# driver.quit()
